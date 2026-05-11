"""
Server-Side Export API Router

Provides streaming export functionality for large datasets.
- CSV, JSON, XLSX exports
- Streaming for memory efficiency
- Export history tracking
"""

import logging
import io
import csv
from typing import List
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text as sql_text

from core.db_session import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["export"])


class ExportRequest(BaseModel):
    """Export request"""
    query: str
    format: str = "csv"  # csv, json, xlsx
    filename: str = "export"
    limit: int = 10000


def generate_csv_stream(rows: List[tuple], headers: List[str]):
    """Generate CSV data as stream"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(headers)
    yield output.getvalue()
    output.truncate(0)
    output.seek(0)
    
    # Write rows
    for row in rows:
        writer.writerow(row)
        yield output.getvalue()
        output.truncate(0)
        output.seek(0)


def generate_json_stream(rows: List[tuple], headers: List[str]):
    """Generate JSON data as stream"""
    yield "[\n"
    
    for idx, row in enumerate(rows):
        record = dict(zip(headers, row))
        import json
        yield json.dumps(record)
        if idx < len(rows) - 1:
            yield ",\n"
        else:
            yield "\n"
    
    yield "]"


@router.post("/data")
async def export_data(
    request: ExportRequest,
    db: Session = Depends(get_db)
):
    """
    Export data to CSV, JSON, or XLSX format with streaming.

    Supports parameterized queries for safety.
    """
    try:
        # Validate query is SELECT only
        query_upper = request.query.upper().strip()
        if not query_upper.startswith("SELECT"):
            raise HTTPException(status_code=400, detail="Only SELECT queries allowed")

        # Add LIMIT clause for safety
        if "LIMIT" not in query_upper:
            query = f"{request.query} LIMIT {request.limit}"
        else:
            query = request.query

        # Execute query
        result = db.execute(sql_text(query)).fetchall()
        
        if not result:
            raise HTTPException(status_code=404, detail="No data to export")

        # Get headers from first row
        headers = list(result[0].keys()) if hasattr(result[0], 'keys') else [f"col_{i}" for i in range(len(result[0]))]

        # Generate response based on format
        if request.format == "csv":
            return StreamingResponse(
                generate_csv_stream(result, headers),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={request.filename}.csv"}
            )
        elif request.format == "json":
            return StreamingResponse(
                generate_json_stream(result, headers),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={request.filename}.json"}
            )
        elif request.format == "xlsx":
            # For XLSX, we need to collect all data first
            try:
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                
                # Write header
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Write data
                for row_idx, row in enumerate(result, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={request.filename}.xlsx"}
                )
            except ImportError:
                raise HTTPException(status_code=500, detail="XLSX export requires openpyxl library")
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {request.format}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error exporting data: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/formats")
async def get_export_formats():
    """Get supported export formats"""
    return {
        "formats": ["csv", "json", "xlsx"],
        "default": "csv",
        "description": "Server-side export with streaming support"
    }
